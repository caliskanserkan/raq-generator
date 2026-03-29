import streamlit as st
import pandas as pd
import datetime, io, os
from pathlib import Path

st.set_page_config(
    page_title="RAQ Form Generator – REC Havacilik",
    page_icon="✈",
    layout="centered"
)

# ── Custom CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main { background-color: #1A252F; }
    .block-container { padding-top: 1rem; }
    .banner {
        background-color: #C0392B;
        padding: 18px 24px;
        border-radius: 8px;
        margin-bottom: 20px;
    }
    .banner h1 { color: white; margin: 0; font-size: 24px; }
    .banner p  { color: #FADBD8; margin: 0; font-size: 13px; }
    .info-card {
        background-color: #2C3E50;
        padding: 14px 20px;
        border-radius: 8px;
        margin: 10px 0;
    }
    .info-card h3 { color: white; margin: 0 0 4px 0; font-size: 16px; }
    .info-card p  { color: #AEB6BF; margin: 0; font-size: 12px; }
    .risk-box {
        background-color: #FADBD8;
        border: 2px solid #C0392B;
        border-radius: 6px;
        padding: 12px 16px;
        margin: 8px 0;
    }
    .risk-box p { color: #C0392B; font-weight: bold; margin: 0; font-size: 13px; }
    stButton > button {
        background-color: #1E8449;
        color: white;
        font-size: 16px;
        font-weight: bold;
        border-radius: 8px;
        padding: 12px 32px;
        border: none;
        width: 100%;
    }
</style>
""", unsafe_allow_html=True)

# ── Database ──────────────────────────────────────────────────────────────────
@st.cache_data
def load_db():
    try:
        import openpyxl
        wb = openpyxl.load_workbook("database.xlsx", data_only=True)

        airports = {}
        ws = wb['AIRPORT_DB']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                airports[str(row[0]).upper()] = {
                    "icao":     str(row[0]),
                    "name":     str(row[1]) if row[1] else str(row[0]),
                    "section1": str(row[2]) if row[2] else "",
                    "section2": str(row[3]) if row[3] else "",
                    "section3": str(row[4]) if row[4] else "",
                    "updated":  str(row[6])[:10] if row[6] else "",
                    "category": str(row[8]) if row[8] else "A",
                }

        risks = {}
        ws2 = wb['MATRIX_STORE']
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if row[0]:
                s = [v for v in row[1:11]  if isinstance(v, (int, float))]
                l = [v for v in row[11:21] if isinstance(v, (int, float))]
                cat = str(row[27]) if row[27] else "A"
                mit = str(row[24]) if row[24] else ""
                if s and l:
                    sc = round(sum(sorted(s,reverse=True)[:3])/3 +
                               sum(sorted(l,reverse=True)[:3])/3 - 1)
                    if cat == "C": sc += 1
                    if sc <= 6:    rl, ops = "LOW",    "DISPATCH OK"
                    elif sc <= 9:  rl, ops = "MEDIUM",  "DISPATCH OK"
                    elif sc <= 12:
                        rl  = "HIGH"
                        ops = "OPS MANAGER APPROVAL REQUIRED" if cat=="C" else "CAPTAIN REVIEW / DISPATCH COORDINATION"
                    else:          rl, ops = "EXTREME","OPS MANAGER APPROVAL REQUIRED"
                else:
                    rl, ops = "N/A", "N/A"
                risks[str(row[0]).upper()] = {
                    "risk_level":   rl,
                    "ops_approval": ops,
                    "mitigation":   mit,
                }
        return airports, risks
    except Exception as e:
        st.error(f"Veritabani yuklenemedi: {e}")
        return {}, {}

# ── PDF Generator ─────────────────────────────────────────────────────────────
def generate_pdf(airport, risk, flight_info):
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os

    # Register DejaVu fonts for full Turkish/Unicode support
    font_paths = [
        ('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',         'DejaVu'),
        ('/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',    'DejaVu-Bold'),
        ('/usr/share/fonts/truetype/dejavu/DejaVuSans-Oblique.ttf', 'DejaVu-Oblique'),
    ]
    for path, name in font_paths:
        if os.path.exists(path) and name not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(TTFont(name, path))

    # Use DejaVu if available, else fallback to Helvetica
    def F(style=''):
        registered = pdfmetrics.getRegisteredFontNames()
        if style == 'Bold'    and 'DejaVu-Bold'    in registered: return 'DejaVu-Bold'
        if style == 'Oblique' and 'DejaVu-Oblique' in registered: return 'DejaVu-Oblique'
        if 'DejaVu' in registered: return 'DejaVu'
        if style == 'Bold':    return F('Bold')
        if style == 'Oblique': return F('Oblique')
        return 'Helvetica' 

    RED   = colors.HexColor("#C0392B")
    DARK  = colors.HexColor("#1A252F")
    MID   = colors.HexColor("#2C3E50")
    STEEL = colors.HexColor("#4A6274")
    LIGHT = colors.HexColor("#FDFEFE")
    ALT   = colors.HexColor("#F2F3F4")
    INPB  = colors.HexColor("#EBF5FB")
    RISKB = colors.HexColor("#FADBD8")
    BORD  = colors.HexColor("#AEB6BF")
    WHITE = colors.white

    PW, PH = A4
    ML = MR = 15 * mm
    W  = PW - ML - MR

    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=A4)
    c.setTitle(f"RAQ Form – {airport.get('icao','')} – {flight_info.get('date','')}")
    form = c.acroForm

    y = PH - 12*mm

    def rect(x, yt, w, h, fill=None, stroke=BORD, sw=0.5):
        c.saveState()
        c.setLineWidth(sw)
        c.setStrokeColor(stroke)
        if fill:
            c.setFillColor(fill)
            c.rect(x, yt-h, w, h, fill=1, stroke=1)
        else:
            c.rect(x, yt-h, w, h, fill=0, stroke=1)
        c.restoreState()

    def txt(text, x, yb, font='Helvetica', size=9, color=DARK,
            align='left', max_w=None):
        c.saveState()
        c.setFont(font, size)
        c.setFillColor(color)
        if align == 'center' and max_w:
            x = x + (max_w - c.stringWidth(text, font, size)) / 2
        elif align == 'right' and max_w:
            x = x + max_w - c.stringWidth(text, font, size)
        c.drawString(x, yb, text)
        c.restoreState()

    def wrap(text, font, size, max_w):
        import re
        words = re.split(r'(\s+)', text)
        lines, line = [], ""
        for word in words:
            test = line + word
            if c.stringWidth(test, font, size) <= max_w:
                line = test
            else:
                if line.strip(): lines.append(line.strip())
                line = word.lstrip()
        if line.strip(): lines.append(line.strip())
        return lines

    def sec(yt, text, h=17):
        rect(ML, yt, W, h, fill=MID, stroke=DARK, sw=1)
        txt(f"  {text}", ML+4, yt-h+4, F('Bold'), 9, WHITE)
        return yt - h

    def subhdr(yt, text, h=13):
        rect(ML, yt, W, h, fill=STEEL, stroke=STEEL)
        txt(f"  {text}", ML+6, yt-h+3, F('Bold'), 8, WHITE)
        return yt - h

    def textblock(yt, text, bg=INPB, pad=8):
        clean = text.replace('\\n', '\n')
        parts = [p.strip() for p in clean.split('•') if p.strip()]
        lines = []
        for p in parts:
            lines.extend(wrap(f"• {p}", F(), 8.5, W-pad*2-4))
        if not parts:
            lines = wrap(clean, F(), 8.5, W-pad*2-4)
        h = len(lines)*11 + pad*2
        rect(ML, yt, W, h, fill=bg, stroke=BORD)
        ty = yt - pad - 8
        for line in lines:
            txt(line, ML+pad+2, ty, F(), 8.5, DARK)
            ty -= 11
        return yt - h

    def cb_row(yt, name, label, bg=LIGHT, h=18, sz=10):
        rect(ML, yt, W, h, fill=bg, stroke=BORD)
        cb_y = yt - h + (h-sz)/2
        form.checkbox(name=name, tooltip=label,
                      x=ML+6, y=cb_y, size=sz,
                      checked=False, buttonStyle='check',
                      borderColor=DARK, fillColor=WHITE,
                      textColor=RED, forceBorder=True)
        txt(label, ML+6+sz+6, yt-h/2-3.5, F(), 9, DARK)
        return yt - h

    def col_hdrs(yt, cols, h=15):
        x = ML
        for label, w in cols:
            rect(x, yt, w, h, fill=STEEL, stroke=BORD)
            txt(label, x, yt-h+4, F('Bold'), 8.5, WHITE, 'center', w)
            x += w
        return yt - h

    def col_data(yt, cols, h=21):
        x = ML
        for text, w, font, size, color, align in cols:
            rect(x, yt, w, h, fill=INPB, stroke=BORD)
            txt(text, x, yt-h+5, font, size, color, align, w)
            x += w
        return yt - h

    # ── Header ────────────────────────────────────────────────────────────────
    hh = 26
    rect(ML,          y, W*0.72, hh, fill=RED,  stroke=DARK, sw=1)
    rect(ML + W*0.72, y, W*0.28, hh, fill=MID,  stroke=DARK, sw=1)
    txt("REC HAVACILIK  ✈  YOL VE MEYDAN YETERLİLİK EĞİTİM FORMU",
        ML+8, y-hh+8, F('Bold'), 10, WHITE)
    txt("FOP-FRM02 | Rev 0 | 2023-10-31",
        ML+W*0.72+4, y-hh+8, F('Bold'), 7, WHITE)
    y -= hh
    sh = 13
    rect(ML, y, W, sh, fill=LIGHT, stroke=BORD)
    txt("ROUTE AND AERODROME QUALIFICATION TRAINING FORM",
        ML, y-sh+3, F('Oblique'), 8, STEEL, 'center', W)
    y -= sh + 3

    # ── Flight Details ────────────────────────────────────────────────────────
    y = sec(y, "FLIGHT DETAILS")
    cw = [W*0.18, W*0.15, W*0.335, W*0.335]
    y = col_hdrs(y, [("Date",cw[0]),("A/C Type",cw[1]),("PIC",cw[2]),("SIC",cw[3])])
    y = col_data(y, [
        (flight_info.get("date",""),    cw[0], F('Bold'), 9, DARK, 'center'),
        (flight_info.get("ac_type",""), cw[1], F('Bold'), 9, DARK, 'center'),
        (flight_info.get("pic",""),     cw[2], F('Bold'), 9, DARK, 'center'),
        (flight_info.get("sic",""),     cw[3], F('Bold'), 9, DARK, 'center'),
    ])
    y -= 3

    # ── Aerodrome ─────────────────────────────────────────────────────────────
    y = sec(y, "AERODROME")
    cw2 = [W*0.50, W*0.25, W*0.25]
    y = col_hdrs(y, [("Airport Name",cw2[0]),("ICAO",cw2[1]),("Category",cw2[2])])
    ah = 22; x = ML
    for text, w, font, size, color in [
        (airport.get("name",""),     cw2[0], F('Bold'), 9,  DARK),
        (airport.get("icao",""),     cw2[1], F('Bold'), 9,  DARK),
        (airport.get("category",""), cw2[2], F('Bold'), 13, RED),
    ]:
        rect(x, y, w, ah, fill=LIGHT, stroke=BORD)
        txt(text, x, y-ah+5, font, size, color, 'center', w)
        x += w
    y -= ah + 3

    # ── Familiarization ───────────────────────────────────────────────────────
    y = sec(y, "FAMILIARIZATION CONDUCTED BY")
    hw = W/2
    y = col_hdrs(y, [("Self Briefing",hw),("Local Authority",hw)])
    fh = 20
    rect(ML,    y, hw, fh, fill=LIGHT, stroke=BORD)
    rect(ML+hw, y, hw, fh, fill=LIGHT, stroke=BORD)
    csz = 12
    cby = y - fh + (fh-csz)/2
    form.checkbox(name='fam_self',  tooltip='Self Briefing',
                  x=ML+hw/2-csz/2, y=cby, size=csz,
                  checked=False, buttonStyle='check',
                  borderColor=DARK, fillColor=WHITE, textColor=RED, forceBorder=True)
    form.checkbox(name='fam_local', tooltip='Local Authority',
                  x=ML+hw+hw/2-csz/2, y=cby, size=csz,
                  checked=False, buttonStyle='check',
                  borderColor=DARK, fillColor=WHITE, textColor=RED, forceBorder=True)
    y -= fh + 3

    # ── Briefing Items ────────────────────────────────────────────────────────
    y = sec(y, "FOLLOWING ITEMS WERE BRIEFED AND FAMILIARIZED FOR THE ROUTE FLOWN")
    for i, (name, label) in enumerate([
        ("cb_terrain",    "Terrain and Safe Altitudes"),
        ("cb_comms",      "Communication and ATC Facilities"),
        ("cb_sar",        "Search & Rescue Procedures"),
        ("cb_layout",     "Airport Layout"),
        ("cb_approach",   "Approach Aids"),
        ("cb_instrument", "Instrument Approach and Hold Procedures"),
        ("cb_minima",     "Operating Minima"),
    ]):
        y = cb_row(y, name, label, bg=LIGHT if i%2==0 else ALT)
    y -= 3

    # ── Special Items ─────────────────────────────────────────────────────────
    y = sec(y, "SPECIAL ITEMS BRIEFED DUE TO AERODROME CATEGORY")
    for i, (name, label) in enumerate([
        ("sp_1","(1) Non-standard approach aids or approach patterns"),
        ("sp_2","(2) Unusual local weather conditions"),
        ("sp_3","(3) Unusual characteristics or performance limitations"),
        ("sp_4","(4) Other relevant considerations: obstructions, physical layout, lighting etc."),
        ("sp_5","(5) Category C aerodromes: additional considerations for approach / landing / take-off."),
    ]):
        y = cb_row(y, name, label, bg=LIGHT if i%2==0 else ALT)
    y -= 3

    # ── PPS Sections ──────────────────────────────────────────────────────────
    y = sec(y, "SPECIAL REMARKS  –  PPS BRIEFING  (AUTO FROM DATABASE)")
    for label, key in [
        ("SECTION 1  –  Traffic / ATC / Taxi / Runway Ops", "section1"),
        ("SECTION 2  –  Meteorology / Wind",                "section2"),
        ("SECTION 3  –  Security / Handling / Navigation",  "section3"),
    ]:
        y = subhdr(y, label)
        y = textblock(y, airport.get(key,"N/A"))
        y -= 2
    y -= 3

    # ── Risk Summary ──────────────────────────────────────────────────────────
    y = sec(y, "AERODROME RISK SUMMARY  (AUTO – DATABASE)")
    if risk:
        rt = (f"RISK LEVEL: {risk['risk_level']}   |   CAT: {airport.get('category','')}   "
              f"|   {risk['ops_approval']}\nMITIGATION: {risk['mitigation']}")
    else:
        rt = "Risk verisi bulunamadi."
    y = textblock(y, rt, bg=RISKB, pad=8)
    y -= 3

    # ── Certification ─────────────────────────────────────────────────────────
    cert = ("I hereby certify that route and aerodrome familiarization was completed "
            "for the flight in accordance with AMC1 ORO.FC.105 b(2);c and OM PART C.")
    ch = 24
    rect(ML, y, W, ch, fill=LIGHT, stroke=BORD)
    lines = wrap(cert, F('Oblique'), 8, W-16)
    ty = y - 6
    for line in lines:
        txt(line, ML+8, ty, F('Oblique'), 8, STEEL)
        ty -= 10
    y -= ch

    today = datetime.date.today().strftime("%Y-%m-%d")
    cw3 = [W*0.22, W*0.55, W*0.23]
    x = ML
    for text, w, font, size, color, align in [
        ("Completed by:", cw3[0], F('Bold'), 9, DARK,  'center'),
        (flight_info.get("pic",""), cw3[1], F('Bold'), 9, DARK, 'center'),
        (today,          cw3[2], F(),       8, STEEL, 'right'),
    ]:
        rect(x, y, w, 20, fill=LIGHT, stroke=BORD)
        txt(text, x, y-15, font, size, color, align, w)
        x += w

    c.save()
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
#  STREAMLIT UI
# ══════════════════════════════════════════════════════════════════════════════
airports, risks = load_db()

# Banner
st.markdown("""
<div class="banner">
  <h1>✈ REC HAVACILIK – RAQ Form Generator</h1>
  <p>Route and Aerodrome Qualification Training Form</p>
</div>
""", unsafe_allow_html=True)

# ICAO Input
col1, col2 = st.columns([1, 2])
with col1:
    icao = st.text_input("ICAO Kodu", max_chars=4,
                          placeholder="EHAM",
                          label_visibility="visible").upper().strip()

# Airport Info Card
if icao:
    if icao in airports:
        ap = airports[icao]
        rk = risks.get(icao)
        st.markdown(f"""
        <div class="info-card">
          <h3>✔ {ap['name']}</h3>
          <p>ICAO: {ap['icao']}   |   Kategori: {ap['category']}   |   Güncelleme: {ap['updated']}</p>
        </div>
        """, unsafe_allow_html=True)
        if rk:
            st.markdown(f"""
            <div class="risk-box">
              <p>RISK: {rk['risk_level']}  |  {rk['ops_approval']}</p>
            </div>
            """, unsafe_allow_html=True)
    elif len(icao) == 4:
        st.error("⚠ Meydan veritabanında bulunamadı.")

st.divider()

# Flight Info
st.subheader("Uçuş Bilgileri")
col1, col2 = st.columns(2)
with col1:
    pic = st.text_input("PIC (Ad Soyad / Kod)")
    date = st.date_input("Uçuş Tarihi", value=datetime.date.today())
with col2:
    sic = st.text_input("SIC (Ad Soyad / Kod)")
    ac  = st.text_input("A/C Type", value="TC-REC")

st.divider()

# Generate Button
if st.button("📄  RAQ FORM PDF OLUŞTUR", use_container_width=True):
    if not icao or icao not in airports:
        st.error("Geçerli bir ICAO kodu girin.")
    elif not pic:
        st.error("PIC bilgisini girin.")
    else:
        ap = airports[icao]
        rk = risks.get(icao)
        fi = {
            "date":    date.strftime("%Y-%m-%d"),
            "ac_type": ac,
            "pic":     pic,
            "sic":     sic,
        }
        with st.spinner("PDF oluşturuluyor..."):
            try:
                pdf_bytes = generate_pdf(ap, rk, fi)
                fname = f"RAQ_{icao}_{date.strftime('%Y-%m-%d')}.pdf"
                st.success(f"✔ PDF hazır: {fname}")
                st.download_button(
                    label="⬇  PDF İndir",
                    data=pdf_bytes,
                    file_name=fname,
                    mime="application/pdf",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"Hata: {e}")

# Footer
st.markdown("---")
st.caption("© REC HAVACILIK  –  AMC1 ORO.FC.105 b(2);c")
